#エクセルデータ使う用のやつ
import openpyxl

#エクセルデータの指定
workbook = openpyxl.load_workbook("meal.xlsx")

#シート番号を指定
sheet = workbook["Sheet1"]

#辞書型を宣言
suppliers = {}

#辞書型に商品名をぶち込むやつ
for i in range(2,12):
    suppliers[sheet.cell(row = i,column = 1).value] = sheet.cell(row = i, column = 3).value



#辞書型表示
print(suppliers)
